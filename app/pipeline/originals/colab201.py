# -*- coding: utf-8 -*-
"""
Colab: JSON -> Excel 転記（仕様駆動）
  + LibreOfficeで再計算
  + 2番目以降の全シートで「数式→結果（値貼り）」へ置換
  + 追加仕様：勘定科目が空の場合、備考（集計方法）も必ず空にする

前提（/tmp/work/）:
  - エクセル転記仕様.xlsx
  - output_updated.json
  - CF付財務分析表（経営指標あり）_ReadingData.xlsx

出力（/tmp/work/）:
  - CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx
  - transfer_log.txt
"""

import json
import re
import math
import os
import subprocess
from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string


# ==========
# パス設定（既定: /tmp/work/, WORK_DIR で上書き可）
# ==========
BASE_DIR = os.environ.get("WORK_DIR", "/tmp/work")
SPEC_PATH = os.path.join(BASE_DIR, "エクセル転記仕様.xlsx")
SRC_EXCEL_PATH = os.path.join(BASE_DIR, "CF付財務分析表（経営指標あり）_ReadingData.xlsx")
JSON_PATH = os.path.join(BASE_DIR, "output_updated.json")

INTERIM_EXCEL_PATH = os.path.join(BASE_DIR, "_interim_after_transfer.xlsx")
RECALC_EXCEL_PATH = os.path.join(BASE_DIR, "_recalc_by_libreoffice.xlsx")

OUT_EXCEL_PATH = os.path.join(BASE_DIR, "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx")
OUT_LOG_PATH = os.path.join(BASE_DIR, "transfer_log.txt")

TARGET_SHEET_NAME = "財務諸表（入力）"


# ==========
# コマンド実行
# ==========
def run_cmd(cmd: list[str]) -> str:
    p = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
    if p.returncode != 0:
        raise RuntimeError(f"Command failed: {' '.join(cmd)}\n--- output ---\n{p.stdout}")
    return p.stdout


# ==========
# 行集合パース
# ==========
def parse_row_set(expr: str) -> set[int]:
    if expr is None:
        return set()
    s = str(expr).strip()
    if not s:
        return set()
    s = s.replace(" ", "").replace("　", "")
    parts = [p for p in s.split(",") if p]
    rows: set[int] = set()
    for p in parts:
        if "-" in p:
            a, b = p.split("-", 1)
            if a == "" or b == "":
                raise ValueError(f"Invalid range token: {p}")
            a_i = int(a)
            b_i = int(b)
            if a_i > b_i:
                raise ValueError(f"Range start > end: {p}")
            rows.update(range(a_i, b_i + 1))
        else:
            rows.add(int(p))
    return rows


# ==========
# 値変換（JSON -> Excel）
# ==========
def coerce_value(v):
    if v is None:
        return None

    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v

    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return None

        # 数値文字列対応（カンマ除去）
        s2 = s.replace(",", "").replace("，", "")

        # (123) -> -123
        m = re.fullmatch(r"\(([-+]?\d+(\.\d+)?)\)", s2)
        if m:
            s2 = "-" + m.group(1)

        try:
            if re.fullmatch(r"[-+]?\d+", s2):
                return int(s2)
            return float(s2)
        except Exception:
            # 数値化できないものは文字列のまま
            return s

    return v


# ==========
# 結合セル対応
# ==========
def writable_cell(ws, row: int, col_letter: str):
    col_idx = column_index_from_string(col_letter)
    requested_cell = ws.cell(row=row, column=col_idx)
    coord = f"{col_letter}{row}"

    if not ws.merged_cells.ranges:
        return requested_cell

    for mr in ws.merged_cells.ranges:
        if coord in mr:
            min_col, min_row, max_col, max_row = mr.bounds
            return ws.cell(row=min_row, column=min_col)

    return requested_cell


# ==========
# Excel起動時フル再計算設定（保険）
# ==========
def set_recalc_on_load(workbook: openpyxl.Workbook):
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True


# ==========
# LibreOfficeで再計算（ODS変換往復）
# ==========
def libreoffice_recalc_via_convert(src_xlsx: str, out_xlsx: str, workdir="/tmp/work"):

    tmp_ods = os.path.join(workdir, "_tmp_recalc.ods")
    if os.path.exists(tmp_ods):
        os.remove(tmp_ods)

    # xlsx -> ods
    run_cmd([
        "bash", "-lc",
        f'libreoffice --headless --nologo --nodefault --nofirststartwizard '
        f'--convert-to ods --outdir "{workdir}" "{src_xlsx}"'
    ])

    base = os.path.splitext(os.path.basename(src_xlsx))[0]
    generated_ods = os.path.join(workdir, base + ".ods")
    if os.path.exists(generated_ods):
        os.replace(generated_ods, tmp_ods)
    elif not os.path.exists(tmp_ods):
        ods_candidates = [f for f in os.listdir(workdir) if f.lower().endswith(".ods")]
        if len(ods_candidates) == 1:
            os.replace(os.path.join(workdir, ods_candidates[0]), tmp_ods)
        else:
            raise RuntimeError("LibreOffice conversion to ODS did not produce a single identifiable .ods file.")

    # ods -> xlsx
    if os.path.exists(out_xlsx):
        os.remove(out_xlsx)

    run_cmd([
        "bash", "-lc",
        f'libreoffice --headless --nologo --nodefault --nofirststartwizard '
        f'--convert-to xlsx --outdir "{workdir}" "{tmp_ods}"'
    ])

    generated_xlsx = os.path.join(workdir, os.path.splitext(os.path.basename(tmp_ods))[0] + ".xlsx")
    if not os.path.exists(generated_xlsx):
        xlsx_candidates = [f for f in os.listdir(workdir) if f.lower().endswith(".xlsx")]
        tmp_candidates = [f for f in xlsx_candidates if "_tmp_recalc" in f]
        if len(tmp_candidates) == 1:
            generated_xlsx = os.path.join(workdir, tmp_candidates[0])
        else:
            raise RuntimeError("LibreOffice conversion back to XLSX did not produce an identifiable .xlsx file.")

    os.replace(generated_xlsx, out_xlsx)

    if os.path.exists(tmp_ods):
        os.remove(tmp_ods)


# ==========
# 数式→値貼り（2番目以降のシート）
# ==========
def replace_formulas_with_values_from_cache(wb_formula: openpyxl.Workbook, wb_values: openpyxl.Workbook):
    sheetnames = wb_formula.sheetnames
    for si in range(1, len(sheetnames)):  # 2番目以降
        name = sheetnames[si]
        ws_f = wb_formula[name]
        ws_v = wb_values[name]

        max_r = ws_f.max_row
        max_c = ws_f.max_column

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell_f = ws_f.cell(row=r, column=c)
                is_formula = (cell_f.data_type == "f") or (isinstance(cell_f.value, str) and cell_f.value.startswith("="))
                if not is_formula:
                    continue

                val = ws_v.cell(row=r, column=c).value
                col_letter = openpyxl.utils.get_column_letter(c)
                target_cell = writable_cell(ws_f, r, col_letter)
                target_cell.value = val


# ==========
# 仕様（転記ルール）読み取り：ヘッダ自動検出
# ==========
def find_header_row_and_cols(ws):
    """
    先頭100行を走査し、次のヘッダを含む行を探す。
      必須: 対象, 転記列, 転記行
      任意: キー名
    戻り値:
      (header_row, colmap) or (None, None)
      colmap: {"対象": col, "転記列": col, "転記行": col, "キー名": col(あれば)}
    """
    max_scan = min(ws.max_row, 100)
    for r in range(1, max_scan + 1):
        colmap = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s in ("対象", "キー名", "転記列", "転記行"):
                colmap[s] = c
        if ("対象" in colmap) and ("転記列" in colmap) and ("転記行" in colmap):
            return r, colmap
    return None, None


def load_rules(spec_path: str):
    """
    rules 要素:
      { target, key_name, column, allowed_rows }
    """
    spec_wb = openpyxl.load_workbook(spec_path, data_only=True)

    # 優先順（よくある名前から）
    sheet_order = []
    for n in ("ルール(正)", "Sheet2", "Sheet1"):
        if n in spec_wb.sheetnames:
            sheet_order.append(n)
    for n in spec_wb.sheetnames:
        if n not in sheet_order:
            sheet_order.append(n)

    for sheet_name in sheet_order:
        ws = spec_wb[sheet_name]
        header_row, colmap = find_header_row_and_cols(ws)
        if header_row is None:
            continue

        has_key = "キー名" in colmap

        # 3列形式の場合は対象からキー名を補完
        target_to_key = {
            "区分": "category",
            "勘定科目": "account_name",
            "前々期": "value_t-2",
            "前期": "value_t-1",
            "今期": "value_t",
            "備考": "集計方法",
        }

        rules = []
        for r in range(header_row + 1, ws.max_row + 1):
            target = ws.cell(r, colmap["対象"]).value
            col = ws.cell(r, colmap["転記列"]).value
            row_expr = ws.cell(r, colmap["転記行"]).value
            key_name = ws.cell(r, colmap["キー名"]).value if has_key else None

            # 空行・注釈行をスキップ
            if target is None and col is None and row_expr is None and (not has_key or key_name is None):
                continue
            if target is None or col is None or row_expr is None:
                continue

            target_s = str(target).strip()
            col_s = str(col).strip()
            allowed_rows = parse_row_set(str(row_expr))

            if has_key:
                if key_name is None:
                    continue
                key_s = str(key_name).strip()
            else:
                key_s = target_to_key.get(target_s)
                if key_s is None:
                    continue

            rules.append({
                "target": target_s,
                "key_name": key_s,
                "column": col_s,
                "allowed_rows": allowed_rows,
            })

        if rules:
            return rules, sheet_name, header_row

    raise ValueError("転記ルールが見つかりません。仕様シートに '対象/転記列/転記行' のヘッダ行が必要です。")


# ==========
# メイン
# ==========
def main():
    for p in [SPEC_PATH, SRC_EXCEL_PATH, JSON_PATH]:
        if not os.path.exists(p):
            raise FileNotFoundError(f"Required file not found: {p}")

    rules, rules_sheet, header_row = load_rules(SPEC_PATH)

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        records = json.load(f)
    if not isinstance(records, list):
        raise ValueError("output_updated.json must be a JSON array (list) of objects.")

    wb = openpyxl.load_workbook(SRC_EXCEL_PATH)
    if TARGET_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Sheet "{TARGET_SHEET_NAME}" not found in {SRC_EXCEL_PATH}')
    ws = wb[TARGET_SHEET_NAME]

    set_recalc_on_load(wb)

    # 仕様キー名 -> JSON実キー名（英語仕様→日本語JSON）
    SPEC_TO_JSON_KEY = {
        "category": "区分",
        "account_name": "勘定科目",
        "value_t-2": "前々期",
        "value_t-1": "前期",
        "value_t": "今期",
        # 備考（集計方法）別名も許可
        "remark": "集計方法",
        "aggregation_method": "集計方法",
        # '集計方法' は mapping なしでも一致する
    }

    log_lines = []
    log_lines.append("=== Excel Transfer Log ===")
    log_lines.append(f"Timestamp: {datetime.now().isoformat(timespec='seconds')}")
    log_lines.append(f"Spec: {SPEC_PATH}")
    log_lines.append(f"Spec rules sheet: {rules_sheet} (header_row={header_row})")
    log_lines.append(f"Source Excel: {SRC_EXCEL_PATH}")
    log_lines.append(f"JSON: {JSON_PATH}")
    log_lines.append(f"Target Sheet: {TARGET_SHEET_NAME}")
    log_lines.append("")

    stats = {
        "records_total": len(records),
        "records_used": 0,
        "rows_written": 0,
        "cells_written": 0,
        "skipped_sheet_mismatch": 0,
        "skipped_no_cell": 0,
        "skipped_bad_cell": 0,
        "skipped_not_allowed": 0,
        "missing_key": 0,
        "writes_to_merged_anchor": 0,
        "remark_forced_blank": 0,  # 追加仕様で備考を空にした回数
    }

    field_stats = {rule["key_name"]: {"written": 0, "skipped_not_allowed": 0, "missing_value": 0} for rule in rules}

    for rec in records:
        if not isinstance(rec, dict):
            continue

        rec_sheet = rec.get("シート名")
        if rec_sheet is not None and rec_sheet != TARGET_SHEET_NAME:
            stats["skipped_sheet_mismatch"] += 1
            continue

        cell_row_raw = rec.get("セル")
        if cell_row_raw is None:
            stats["skipped_no_cell"] += 1
            continue

        try:
            excel_row = int(str(cell_row_raw).strip())
        except Exception:
            stats["skipped_bad_cell"] += 1
            continue

        # 追加仕様判定に必要：勘定科目
        account_val = coerce_value(rec.get("勘定科目"))
        account_is_blank = (account_val is None) or (isinstance(account_val, str) and account_val.strip() == "")

        stats["records_used"] += 1
        wrote_any = False

        for rule in rules:
            key_name = rule["key_name"]
            col_letter = rule["column"]
            allowed_rows = rule["allowed_rows"]

            if excel_row not in allowed_rows:
                stats["skipped_not_allowed"] += 1
                field_stats[key_name]["skipped_not_allowed"] += 1
                continue

            json_key = SPEC_TO_JSON_KEY.get(key_name, key_name)

            # ★追加仕様：勘定科目が空なら、備考（集計方法）を必ず空にする
            if json_key == "集計方法" and account_is_blank:
                value = None
                stats["remark_forced_blank"] += 1
            else:
                if json_key not in rec:
                    stats["missing_key"] += 1
                    continue
                value = coerce_value(rec.get(json_key))

            requested_coord = f"{col_letter}{excel_row}"
            cell = writable_cell(ws, excel_row, col_letter)
            if cell.coordinate != requested_coord:
                stats["writes_to_merged_anchor"] += 1

            cell.value = value

            field_stats[key_name]["written"] += 1
            if value is None:
                field_stats[key_name]["missing_value"] += 1

            stats["cells_written"] += 1
            wrote_any = True

        if wrote_any:
            stats["rows_written"] += 1

    wb.save(INTERIM_EXCEL_PATH)

    libreoffice_recalc_via_convert(INTERIM_EXCEL_PATH, RECALC_EXCEL_PATH, workdir="/tmp/work")

    wb_formula = openpyxl.load_workbook(RECALC_EXCEL_PATH, data_only=False)
    wb_values = openpyxl.load_workbook(RECALC_EXCEL_PATH, data_only=True)

    replace_formulas_with_values_from_cache(wb_formula, wb_values)
    set_recalc_on_load(wb_formula)

    wb_formula.save(OUT_EXCEL_PATH)

    log_lines.append("=== Summary ===")
    for k, v in stats.items():
        log_lines.append(f"{k}: {v}")
    log_lines.append("")
    log_lines.append("=== Per Field ===")
    for k, v in field_stats.items():
        log_lines.append(
            f"[{k}] written={v['written']} skipped_not_allowed={v['skipped_not_allowed']} missing_value={v['missing_value']}"
        )
    log_lines.append("")
    log_lines.append("=== Formula->Value Replace ===")
    log_lines.append("Target: sheets with index>=1 (2nd and later sheets)")
    log_lines.append(f"Recalc file: {RECALC_EXCEL_PATH}")
    log_lines.append(f"Final output: {OUT_EXCEL_PATH}")
    log_lines.append("")

    with open(OUT_LOG_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))

    print("DONE")
    print("Final Excel:", OUT_EXCEL_PATH)
    print("Log:", OUT_LOG_PATH)
    print("Stats:", stats)


if __name__ == "__main__":
    main()
