def main():
    # ==========================================
    # Colab: JSON -> Excel 転記（仕様駆動）
    #   + LibreOfficeで再計算
    #   + 2番目以降の全シートで「数式→結果（値貼り）」へ置換
    #
    # 前提ファイル（/tmp/work/）:
    #   - エクセル転記仕様.xlsx（Sheet: ルール(正)）
    #   - output_updated.json
    #   - CF付財務分析表（経営指標あり）_ReadingData.xlsx
    #
    # 出力（/tmp/work/）:
    #   - CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx  （最終成果物）
    #   - transfer_log.txt
    # ==========================================
    
    import json
    import re
    import math
    import os
    import subprocess
    from datetime import datetime
    
    import openpyxl
    from openpyxl.utils import column_index_from_string
    
    # ==========
    # パス設定（既定: /tmp/work/, WORK_DIR で上書き可）
    # ==========
    BASE_DIR = os.environ.get("WORK_DIR", "/tmp/work")
    SPEC_PATH = os.path.join(BASE_DIR, "エクセル転記仕様.xlsx")
    SRC_EXCEL_PATH = os.path.join(BASE_DIR, "CF付財務分析表（経営指標あり）_ReadingData.xlsx")
    JSON_PATH = os.path.join(BASE_DIR, "output_updated.json")
    
    # 中間ファイル（転記直後）
    INTERIM_EXCEL_PATH = os.path.join(BASE_DIR, "_interim_after_transfer.xlsx")
    
    # LibreOfficeで再計算させたファイル（数式は残るが、計算結果キャッシュが更新されることを期待）
    RECALC_EXCEL_PATH = os.path.join(BASE_DIR, "_recalc_by_libreoffice.xlsx")
    
    # 最終出力（数式→値貼り後）
    OUT_EXCEL_PATH = os.path.join(BASE_DIR, "CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx")
    OUT_LOG_PATH = os.path.join(BASE_DIR, "transfer_log.txt")
    
    TARGET_SHEET_NAME = "財務諸表（入力）"
    
    
    # ==========
    # 汎用: コマンド実行（失敗時は例外）
    # ==========
    def run_cmd(cmd: list[str]):
        p = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
        if p.returncode != 0:
            raise RuntimeError(f"Command failed: {' '.join(cmd)}\n--- output ---\n{p.stdout}")
        return p.stdout
    
    
    # ==========
    # 転記仕様の「転記行」パース
    # 例: '6-10,12-15,17-26,118,120-121'
    # ==========
    def parse_row_set(expr: str) -> set[int]:
        """
        仕様:
          - a-b は a以上b以下（両端含む）
          - ',' は和集合
          - 空白はトリム
          - 重複は無視（集合）
        """
        if expr is None:
            return set()
        s = str(expr).strip()
        if not s:
            return set()
    
        s = s.replace(" ", "").replace("　", "")
        parts = [p for p in s.split(",") if p]
    
        rows: set[int] = set()
        for p in parts:
            if "-" in p:
                a, b = p.split("-", 1)
                if a == "" or b == "":
                    raise ValueError(f"Invalid range token: {p}")
                a_i = int(a)
                b_i = int(b)
                if a_i > b_i:
                    raise ValueError(f"Range start > end: {p}")
                rows.update(range(a_i, b_i + 1))
            else:
                rows.add(int(p))
        return rows
    
    
    # ==========
    # 値変換（JSON -> Excel）
    # ==========
    def coerce_value(v):
        """
        - None / '' は None（空欄）
        - 数値型はそのまま（NaN/Inf は None）
        - 文字列数値はカンマ除去して int/float へ
        - それ以外は文字列として返す（勘定科目など）
        """
        if v is None:
            return None
    
        if isinstance(v, (int, float)):
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                return None
            return v
    
        if isinstance(v, str):
            s = v.strip()
            if s == "":
                return None
    
            s2 = s.replace(",", "").replace("，", "")
    
            # (123) -> -123
            m = re.fullmatch(r"\(([-+]?\d+(\.\d+)?)\)", s2)
            if m:
                s2 = "-" + m.group(1)
    
            try:
                if re.fullmatch(r"[-+]?\d+", s2):
                    return int(s2)
                return float(s2)
            except Exception:
                return s
    
        return v
    
    
    # ==========
    # 結合セル対応（MergedCell 対策）
    # openpyxl では結合セル範囲の「左上セル」以外は書き込み不可
    # ==========
    def writable_cell(ws, row: int, col_letter: str):
        """
        指定セルが結合セル範囲内なら、その結合範囲の左上（アンカー）セルを返す。
        """
        col_idx = column_index_from_string(col_letter)
        requested_cell = ws.cell(row=row, column=col_idx)
        coord = f"{col_letter}{row}"
    
        if not ws.merged_cells.ranges:
            return requested_cell
    
        for mr in ws.merged_cells.ranges:
            if coord in mr:
                min_col, min_row, max_col, max_row = mr.bounds
                return ws.cell(row=min_row, column=min_col)
    
        return requested_cell
    
    
    # ==========
    # 値更新処理(1): Excel起動時に全再計算させる設定（保険）
    # ==========
    def set_recalc_on_load(workbook: openpyxl.Workbook):
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
    
    
    # ==========
    # LibreOfficeで再計算させた状態のファイルを作る
    # 方針:
    #   1) 中間xlsxを LibreOffice で ODS に変換（ロード時に再計算が走ることを期待）
    #   2) その ODS を LibreOffice で XLSX に変換して戻す
    # これにより「計算結果キャッシュが更新されたxlsx」を得ることを狙う
    # ==========
    def libreoffice_recalc_via_convert(src_xlsx: str, out_xlsx: str, workdir="/tmp/work"):
        # LibreOffice インストール（未インストールなら）
    
        tmp_ods = os.path.join(workdir, "_tmp_recalc.ods")
        # 変換先の自動命名に備えて掃除
        if os.path.exists(tmp_ods):
            os.remove(tmp_ods)
    
        # xlsx -> ods
        # --convert-to は outdir に「同名 + .ods」を作る
        run_cmd([
            "bash", "-lc",
            f'libreoffice --headless --nologo --nodefault --nofirststartwizard '
            f'--convert-to ods --outdir "{workdir}" "{src_xlsx}"'
        ])
    
        # 生成された ods を探す（同名ベースが一般的）
        base = os.path.splitext(os.path.basename(src_xlsx))[0]
        generated_ods = os.path.join(workdir, base + ".ods")
        if os.path.exists(generated_ods):
            os.replace(generated_ods, tmp_ods)
        elif not os.path.exists(tmp_ods):
            # 念のため workdir 内の .ods を探索
            ods_candidates = [f for f in os.listdir(workdir) if f.lower().endswith(".ods")]
            if len(ods_candidates) == 1:
                os.replace(os.path.join(workdir, ods_candidates[0]), tmp_ods)
            else:
                raise RuntimeError("LibreOffice conversion to ODS did not produce a single identifiable .ods file.")
    
        # ods -> xlsx
        # outdir に「_tmp_recalc.xlsx」等ができるので、最後に所望名へリネーム
        if os.path.exists(out_xlsx):
            os.remove(out_xlsx)
    
        run_cmd([
            "bash", "-lc",
            f'libreoffice --headless --nologo --nodefault --nofirststartwizard '
            f'--convert-to xlsx --outdir "{workdir}" "{tmp_ods}"'
        ])
    
        generated_xlsx = os.path.join(workdir, os.path.splitext(os.path.basename(tmp_ods))[0] + ".xlsx")
        if not os.path.exists(generated_xlsx):
            # outdir 内の xlsx を探索
            xlsx_candidates = [f for f in os.listdir(workdir) if f.lower().endswith(".xlsx")]
            # 中間や元ファイルを除きたいので、tmp名を含むものを優先
            tmp_candidates = [f for f in xlsx_candidates if "_tmp_recalc" in f]
            if len(tmp_candidates) == 1:
                generated_xlsx = os.path.join(workdir, tmp_candidates[0])
            else:
                raise RuntimeError("LibreOffice conversion back to XLSX did not produce an identifiable .xlsx file.")
    
        os.replace(generated_xlsx, out_xlsx)
    
        # 後始末
        if os.path.exists(tmp_ods):
            os.remove(tmp_ods)
    
    
    # ==========
    # 2番目以降のシートで「数式→結果（値）」へ置換
    #
    # アプローチ:
    #   - wb_formula : 数式が入っているブック（通常のload_workbook）
    #   - wb_values  : data_only=True で読み込んだ「計算結果（キャッシュ）参照ブック」
    #   - 対象はシート順 index>=1
    #   - 数式セル（cell.data_type == 'f' または cell.valueが"="で始まる）を、
    #     wb_values側の同セルの値で置換する
    # ==========
    def replace_formulas_with_values_from_cache(wb_formula: openpyxl.Workbook, wb_values: openpyxl.Workbook):
        # シート順で 2番目以降（index=1以降）
        sheetnames = wb_formula.sheetnames
        for si in range(1, len(sheetnames)):
            name = sheetnames[si]
            ws_f = wb_formula[name]
            ws_v = wb_values[name]
    
            # 使用範囲（max_row/max_column）で走査
            max_r = ws_f.max_row
            max_c = ws_f.max_column
    
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    cell_f = ws_f.cell(row=r, column=c)
    
                    # 数式判定
                    is_formula = (cell_f.data_type == "f") or (isinstance(cell_f.value, str) and cell_f.value.startswith("="))
                    if not is_formula:
                        continue
    
                    # 値側（計算結果キャッシュ）
                    cell_v = ws_v.cell(row=r, column=c)
                    val = cell_v.value
    
                    # 結合セルの非アンカーに書こうとすると例外になるため、アンカーへ
                    col_letter = openpyxl.utils.get_column_letter(c)
                    target_cell = writable_cell(ws_f, r, col_letter)
                    target_cell.value = val  # None も許容（空欄）
    
            # 置換後も安全のため
            # 数式を消す以上、calc設定は不要だが、残しておいても害はない
    
    
    # ==========
    # 0) 入力存在チェック
    # ==========
    for p in [SPEC_PATH, SRC_EXCEL_PATH, JSON_PATH]:
        if not os.path.exists(p):
            raise FileNotFoundError(f"Required file not found: {p}")
    
    
    # ==========
    # 1) 転記仕様を読み取る
    # ==========
    spec_wb = openpyxl.load_workbook(SPEC_PATH, data_only=True)
    if "ルール(正)" not in spec_wb.sheetnames:
        raise ValueError('Spec workbook must contain sheet "ルール(正)"')
    rules_ws = spec_wb["ルール(正)"]
    
    rules = []
    for r in range(3, rules_ws.max_row + 1):
        target = rules_ws.cell(r, 1).value
        key_name = rules_ws.cell(r, 2).value
        col = rules_ws.cell(r, 3).value
        row_expr = rules_ws.cell(r, 4).value
    
        if target is None and key_name is None and col is None and row_expr is None:
            continue
        if target is None or key_name is None or col is None or row_expr is None:
            continue
    
        allowed_rows = parse_row_set(str(row_expr))
        rules.append({
            "target": str(target),
            "key_name": str(key_name),
            "column": str(col).strip(),
            "allowed_rows": allowed_rows,
        })
    
    if not rules:
        raise ValueError("No rules found in spec sheet 'ルール(正)'.")
    
    
    # ==========
    # 2) JSON を読み取る
    # ==========
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        records = json.load(f)
    if not isinstance(records, list):
        raise ValueError("output_updated.json must be a JSON array (list) of objects.")
    
    
    # ==========
    # 3) 転記先 Excel を読み取る
    # ==========
    wb = openpyxl.load_workbook(SRC_EXCEL_PATH)
    if TARGET_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Sheet "{TARGET_SHEET_NAME}" not found in {SRC_EXCEL_PATH}')
    ws = wb[TARGET_SHEET_NAME]
    
    # 保険で「Excelで開いたら全再計算」もON
    set_recalc_on_load(wb)
    
    # 仕様キー名 -> JSON実キー名（仕様が英語キー、JSONが日本語キーの前提）
    SPEC_TO_JSON_KEY = {
        "category": "区分",
        "account_name": "勘定科目",
        "value_t-2": "前々期",
        "value_t-1": "前期",
        "value_t": "今期",
        # 仕様が日本語キーを使う場合はそのまま一致して拾える
    }
    
    
    # ==========
    # 4) 転記実行（仕様で許可されたセルのみ）
    # ==========
    log_lines = []
    log_lines.append("=== Excel Transfer Log ===")
    log_lines.append(f"Timestamp: {datetime.now().isoformat(timespec='seconds')}")
    log_lines.append(f"Spec: {SPEC_PATH}")
    log_lines.append(f"Source Excel: {SRC_EXCEL_PATH}")
    log_lines.append(f"JSON: {JSON_PATH}")
    log_lines.append(f"Target Sheet: {TARGET_SHEET_NAME}")
    log_lines.append("")
    
    stats = {
        "records_total": len(records),
        "records_used": 0,
        "rows_written": 0,
        "cells_written": 0,
        "skipped_sheet_mismatch": 0,
        "skipped_no_cell": 0,
        "skipped_bad_cell": 0,
        "skipped_not_allowed": 0,
        "missing_key": 0,
        "writes_to_merged_anchor": 0,
    }
    
    field_stats = {rule["key_name"]: {"written": 0, "skipped_not_allowed": 0, "missing_value": 0} for rule in rules}
    
    for rec in records:
        if not isinstance(rec, dict):
            continue
    
        rec_sheet = rec.get("シート名")
        if rec_sheet is not None and rec_sheet != TARGET_SHEET_NAME:
            stats["skipped_sheet_mismatch"] += 1
            continue
    
        cell_row_raw = rec.get("セル")
        if cell_row_raw is None:
            stats["skipped_no_cell"] += 1
            continue
    
        try:
            excel_row = int(str(cell_row_raw).strip())
        except Exception:
            stats["skipped_bad_cell"] += 1
            continue
    
        stats["records_used"] += 1
        wrote_any = False
    
        for rule in rules:
            key_name = rule["key_name"]
            col_letter = rule["column"]
            allowed_rows = rule["allowed_rows"]
    
            if excel_row not in allowed_rows:
                stats["skipped_not_allowed"] += 1
                field_stats[key_name]["skipped_not_allowed"] += 1
                continue
    
            json_key = SPEC_TO_JSON_KEY.get(key_name, key_name)
            if json_key not in rec:
                stats["missing_key"] += 1
                continue
    
            value = coerce_value(rec.get(json_key))
    
            requested_coord = f"{col_letter}{excel_row}"
            cell = writable_cell(ws, excel_row, col_letter)
            if cell.coordinate != requested_coord:
                stats["writes_to_merged_anchor"] += 1
    
            cell.value = value
    
            field_stats[key_name]["written"] += 1
            if value is None:
                field_stats[key_name]["missing_value"] += 1
    
            stats["cells_written"] += 1
            wrote_any = True
    
        if wrote_any:
            stats["rows_written"] += 1
    
    
    # ==========
    # 5) 中間保存（転記後）
    # ==========
    wb.save(INTERIM_EXCEL_PATH)
    
    
    # ==========
    # 6) LibreOfficeで再計算された（と期待する）ファイルを生成
    # ==========
    libreoffice_recalc_via_convert(INTERIM_EXCEL_PATH, RECALC_EXCEL_PATH, workdir="/tmp/work")
    
    
    # ==========
    # 7) 計算結果キャッシュを参照して「数式→値貼り」へ置換
    #   - wb_formula: 再計算後xlsx（数式入りのまま）
    #   - wb_values : 同じファイルを data_only=True で読み、計算結果を取得
    # ==========
    wb_formula = openpyxl.load_workbook(RECALC_EXCEL_PATH, data_only=False)
    wb_values = openpyxl.load_workbook(RECALC_EXCEL_PATH, data_only=True)
    
    replace_formulas_with_values_from_cache(wb_formula, wb_values)
    
    # 念のため計算設定（値貼り後でも問題ない）
    set_recalc_on_load(wb_formula)
    
    # ==========
    # 8) 最終保存
    # ==========
    wb_formula.save(OUT_EXCEL_PATH)
    
    
    # ==========
    # 9) ログ保存
    # ==========
    log_lines.append("=== Summary ===")
    for k, v in stats.items():
        log_lines.append(f"{k}: {v}")
    log_lines.append("")
    log_lines.append("=== Per Field ===")
    for k, v in field_stats.items():
        log_lines.append(
            f"[{k}] written={v['written']} skipped_not_allowed={v['skipped_not_allowed']} missing_value={v['missing_value']}"
        )
    log_lines.append("")
    log_lines.append("=== Formula->Value Replace ===")
    log_lines.append("Target: sheets with index>=1 (2nd and later sheets)")
    log_lines.append(f"Recalc file: {RECALC_EXCEL_PATH}")
    log_lines.append(f"Final output: {OUT_EXCEL_PATH}")
    log_lines.append("")
    
    with open(OUT_LOG_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))
    
    print("DONE")
    print("Final Excel:", OUT_EXCEL_PATH)
    print("Log:", OUT_LOG_PATH)
    print("Stats:", stats)

if __name__ == "__main__":
    main()
